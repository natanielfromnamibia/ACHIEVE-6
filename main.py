"""
ACHIEVE-6: CORE DECISIONING ENGINE 
---------------------------------
AUTHOR: Alfanet Analysis
COMPLIANCE: Bank of Namibia (BoN) Regulation 2025
FUNCTION: Automated NRV Scoring & DTI Risk Assessment
ARCHITECTURE: Offline-First / Decoupled Data Processing
"""
import csv
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load rules
with open("rules.json", "r", encoding="utf-8") as f:
    R = json.load(f)

# Read input.csv 
with open("input.csv", "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    rows = list(reader)
    r = rows[0]  # Processes first row for demo

score = 0

# Income Scoring
for lim, pts in R["scoring"]["income"]:
    if float(r["monthly_income"]) >= lim:
        score += pts
        break

# Employment Years Scoring
for lim, pts in R["scoring"]["years"]:
    if float(r["employment_years"]) >= lim:
        score += pts
        break

# Age Scoring
for lo, hi, pts in R["scoring"]["age"]:
    if lo <= int(r["age"]) <= hi:
        score += pts
        break

# DTI Scoring
for lim, pts in R["scoring"]["dti"]:
    if float(r["debt_to_income"]) <= lim:
        score += pts
        break

# Defaults check
if int(r["defaults_last_24m"]) == 0:
    score += 20

# Color mapping logic
color = R["colors"]["decline"]
if score >= 85: color = R["colors"]["approve"]
elif score >= 65: color = R["colors"]["review"]
elif score >= 45: color = R["colors"]["caution"]

# Save results to Excel
out = "ACHIEVE6_RESULT.xlsx"
wb = load_workbook("input.xlsx")  # open original for formatting
ws = wb.active
ws["H2"] = score  # Final score in Column H

fill = PatternFill("solid", fgColor=color)
for cell in ws[ws.max_row]:
    cell.fill = fill

wb.save(out)
print(f"ACHIEVE-6 → SCORE {score} → {out} READY")
